import io
import html
import json
import re
import time
import requests
import os
import tempfile
import unicodedata
from datetime import datetime, date
from zoneinfo import ZoneInfo

import pandas as pd
import plotly.express as px
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter

from sca_reports import (
    AGENTES_POR_COMUNIDADE,
    resumo_operacional_com_cadastro,
    diagnostico_relatorio,
    gerar_pdf_comunidade,
    gerar_pdf_mensal,
    adicionar_controle_agentes_excel,
    gerar_powerpoint_resumo,
)

APP_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_LIGHT = os.path.join(APP_DIR, "light_logo.png")
LOGO_DEEP_RODAPE = os.path.join(APP_DIR, "desenvolvido_por_deep.png")


# =========================================================
# CONFIGURAÇÃO GERAL
# =========================================================

st.set_page_config(
    page_title="Deep Field | Automação e Inteligência Operacional",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Evita que tradutores automáticos alterem o DOM controlado pelo Streamlit.
# Alterações externas no DOM podem causar NotFoundError em removeChild.
components.html(
    """
    <script>
    try {
        const doc = window.parent.document;
        doc.documentElement.setAttribute('translate', 'no');
        doc.documentElement.classList.add('notranslate');
        const meta = doc.createElement('meta');
        meta.name = 'google';
        meta.content = 'notranslate';
        if (!doc.head.querySelector('meta[name="google"][content="notranslate"]')) {
            doc.head.appendChild(meta);
        }
    } catch (e) { console.debug('Deep Field: proteção de tradução não aplicada', e); }
    </script>
    """,
    height=0,
    width=0,
)

st.markdown(
    """
    <style>
    :root { --sca-verde:#13A88E; --sca-escuro:#075E54; --sca-claro:#E5F6F2; }
    .stApp { background: #F5FAF8; color: #18312E; }
    [data-testid="stSidebar"] { background: linear-gradient(180deg,#075E54 0%,#0B8F7A 55%,#13A88E 100%); }
    [data-testid="stSidebar"] * { color: white; }
    [data-testid="stSidebar"] img { max-width: 210px; margin: 0 auto 8px auto; display:block; }
    .sca-hero { background: linear-gradient(105deg,#075E54,#13A88E,#35BDB1); padding: 24px 28px; border-radius: 18px; color:white; margin-bottom:18px; box-shadow: 0 8px 24px rgba(7,94,84,.18); }
    .sca-hero h1 { color:white; margin:0; font-size:2rem; }
    .sca-hero p { margin:.35rem 0 0 0; opacity:.94; }
    [data-testid="stMetric"] { background:white; border:1px solid #D7EEE9; padding:14px; border-radius:14px; box-shadow:0 4px 14px rgba(7,94,84,.07); }
    .stButton>button, .stDownloadButton>button { border-radius:10px; border:1px solid #0B8F7A; }
    .stButton>button[kind="primary"], .stDownloadButton>button[kind="primary"] { background:#0B8F7A; color:white; }
    h1,h2,h3 { color:#075E54; }

    /* Identidade Deep Field */
    .deep-field-brand {
        font-size: 18px;
        font-weight: 700;
        line-height: 1.3;
        color: #FFFFFF;
        margin: 2px 0 2px 0;
        letter-spacing: 0.2px;
    }
    [data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        padding-bottom: 150px;
    }
    /* Menu lateral sem bolinhas de radio */
    [data-testid="stSidebar"] div[role="radiogroup"] {
        gap: 8px;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label {
        position: relative;
        display: flex;
        align-items: center;
        min-height: 46px;
        padding: 10px 14px !important;
        margin: 0 0 6px 0;
        border-radius: 11px;
        border: 1px solid transparent;
        background: transparent;
        cursor: pointer;
        transition: background-color .18s ease, color .18s ease, transform .18s ease;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label:hover {
        background: rgba(255,255,255,.14);
        transform: translateX(2px);
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label > div:first-child {
        display: none !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label p {
        font-size: 16px !important;
        line-height: 1.35 !important;
        font-weight: 500 !important;
        color: rgba(255,255,255,.94) !important;
        margin: 0 !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label:has(input:checked) {
        background: #E5F6F2 !important;
        border-color: rgba(255,255,255,.65) !important;
        box-shadow: 0 5px 14px rgba(4,68,60,.16);
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label:has(input:checked)::before {
        content: "";
        position: absolute;
        left: 5px;
        top: 9px;
        bottom: 9px;
        width: 4px;
        border-radius: 4px;
        background: #0B8F7A;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] > label:has(input:checked) p {
        color: #075E54 !important;
        font-weight: 700 !important;
    }
    .deep-sidebar-footer {
        position: fixed;
        left: 16px;
        bottom: 14px;
        width: 232px;
        z-index: 999;
        padding: 13px 12px;
        box-sizing: border-box;
        border-top: 1px solid rgba(255,255,255,.24);
        color: #FFFFFF;
        background: transparent;
        text-align: center;
        font-size: 15px;
        line-height: 1.2;
        letter-spacing: .1px;
    }
    .deep-sidebar-footer span { font-weight: 400; opacity: .92; }
    .deep-sidebar-footer strong { font-size: 18px; font-weight: 800; margin-left: 4px; }
    .deep-sidebar-footer .deep-mark {
        display: inline-block;
        width: 16px;
        height: 25px;
        margin-left: 4px;
        vertical-align: middle;
        border: 3px solid #FFFFFF;
        border-left-color: transparent;
        border-radius: 50%;
        box-sizing: border-box;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    if os.path.exists(LOGO_LIGHT):
        st.image(LOGO_LIGHT, use_container_width=True)
    st.markdown('<div class="deep-field-brand notranslate" translate="no">Deep Field</div>', unsafe_allow_html=True)
    st.caption("Automação e Inteligência Operacional")
    st.markdown("---")


# =========================================================
# FUNÇÕES AUXILIARES
# =========================================================

def normalize_text(texto):
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join([c for c in texto if not unicodedata.combining(c)])
    texto = " ".join(texto.split())
    return texto


def agora_sao_paulo():
    return datetime.now(ZoneInfo("America/Sao_Paulo"))


def converter_data_brasil(serie):
    return pd.to_datetime(
        serie,
        errors="coerce",
        dayfirst=True
    ).dt.date


def formatar_data_brasil(valor):
    if pd.isna(valor):
        return ""

    try:
        return pd.to_datetime(valor, dayfirst=True).strftime("%d/%m/%Y")
    except Exception:
        return str(valor)


def encontrar_coluna(df, palavras):
    for col in df.columns:
        col_norm = normalize_text(col)
        if all(normalize_text(p) in col_norm for p in palavras):
            return col
    return None


def safe_sheet_name(nome):
    nome = str(nome)

    for c in ['\\', '/', '*', '[', ']', ':', '?']:
        nome = nome.replace(c, "")

    return nome[:30] if nome else "ABA"


def read_excel_any(uploaded_file, dtype=None):
    uploaded_file.seek(0)
    ext = os.path.splitext(uploaded_file.name)[1].lower()

    if ext == ".xls":
        return pd.read_excel(uploaded_file, dtype=dtype, engine="xlrd")

    return pd.read_excel(uploaded_file, dtype=dtype, engine="openpyxl")


def excel_bytes_from_wb(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def excel_value(valor):
    if isinstance(valor, tuple):
        if len(valor) == 1:
            return valor[0]
        return " - ".join(str(v) for v in valor)

    if pd.isna(valor):
        return ""

    return valor


def aplicar_bordas_e_larguras(ws, max_width=55):
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue

            if cell.value is not None:
                cell.border = borda

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in col_cells:
            if isinstance(cell, MergedCell):
                continue

            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, max_width)


# =========================================================
# BASE DE ENVIO LIGHT V5 - APENAS BACKOFFICE APROVADO
# =========================================================
TIPO = "ADESÃO REALIZADA (morador presente)"
COMUNIDADES = ["COMUNIDADES CORTE OITO", "COMUNIDADES CHACRINHA", "COMUNIDADES SALGUEIRO"]
COLUNAS_ORIGEM = [
    "Code Deep", "Data do registro", "ASRO", *COMUNIDADES,
    "CEP:", "Endereço", "Complemento:", "bairro", "É novo cliente?",
    "NOME COMPLETO:", "CPF:", "RG:", "DATA DE NASCIMENTO:",
    "TELEFONE PARA CONTATO:", "E-MAIL:", "POSSUI NIS?",
    "Informar o número do NIS:", "NÚMERO DO MEDIDOR:",
    "NÚMERO INSTALAÇÃO:", "UNIDADE CONSUMIDORA/ CÓDIGO DO CLIENTE:",
    "FAIXA DE ENQUADRAMENTO ESCOLHIDA PELO CLIENTE:",
    "DATA DE VENCIMENTO DA FATURA:", "FOTO DO RG", "FOTO DO CPF:"
]
ALIASES = {
    "foto rg": "FOTO DO RG", "foto do rg": "FOTO DO RG",
    "foto cpf": "FOTO DO CPF:", "foto do cpf": "FOTO DO CPF:",
}

def normalizar(v):
    s = "" if v is None else str(v)
    s = re.sub(r"_field", "", s, flags=re.I).replace("\xa0", " ").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).casefold().strip(" :")

def localizar_estrutura(conteudo: bytes):
    xls = pd.ExcelFile(io.BytesIO(conteudo), engine="openpyxl")
    for aba in xls.sheet_names:
        previa = pd.read_excel(io.BytesIO(conteudo), sheet_name=aba, header=None, nrows=15, engine="openpyxl")
        for linha in range(len(previa)):
            vals = {normalizar(v) for v in previa.iloc[linha].dropna()}
            if normalizar("Code Deep") in vals and normalizar("Data do registro") in vals:
                return aba, linha
    raise ValueError("Não foi possível localizar a linha de cabeçalho com 'Code Deep' e 'Data do registro'.")

def ler_arquivo(conteudo: bytes):
    aba, header = localizar_estrutura(conteudo)
    df = pd.read_excel(io.BytesIO(conteudo), sheet_name=aba, header=header, dtype=object, engine="openpyxl")
    # O arquivo real possui cabeçalhos repetidos em outros campos. Mapeamos somente nomes necessários.
    canon = {normalizar(c): c for c in COLUNAS_ORIGEM + ["TIPO DE ATENDIMENTO", "Situação Backoffice"]}
    nomes = []
    usados = set()
    for c in df.columns:
        chave = normalizar(c)
        nome = canon.get(chave, ALIASES.get(chave, str(c).strip()))
        if nome in usados:
            nome = str(c).strip()
        usados.add(nome)
        nomes.append(nome)
    df.columns = nomes
    faltantes = [c for c in COLUNAS_ORIGEM + ["TIPO DE ATENDIMENTO", "Situação Backoffice"] if c not in df.columns]
    if faltantes:
        raise ValueError("Colunas necessárias não encontradas: " + ", ".join(faltantes))
    return df, aba, header

def periodo_disponivel(df):
    d = pd.to_datetime(df["Data do registro"], errors="coerce", dayfirst=True).dropna()
    if d.empty:
        raise ValueError("A coluna 'Data do registro' não contém datas válidas.")
    return d.min().date(), d.max().date()

def extrair_urls(valor):
    """Extrai URLs individuais mesmo quando existem duas ou mais na mesma célula."""
    if valor is None or pd.isna(valor):
        return []
    texto = str(valor).replace("\r", "\n")
    return re.findall(r"https?://[^\s]+", texto, flags=re.IGNORECASE)

def analisar_situacao_backoffice(df):
    """Conta os registros aprovados e os que serão excluídos antes da geração."""
    situacoes = df["Situação Backoffice"].fillna("").apply(normalizar)
    aprovados = situacoes.eq("aprovado")
    return {
        "total": int(len(df)),
        "aprovados": int(aprovados.sum()),
        "nao_aprovados": int((~aprovados).sum()),
        "mascara_aprovados": aprovados,
    }

def processar(df, inicio: date, fim: date):
    base = df.copy()
    datas = pd.to_datetime(base["Data do registro"], errors="coerce", dayfirst=True)
    atendimento = base["TIPO DE ATENDIMENTO"].fillna("").astype(str).str.strip()

    # Critério obrigatório e prioritário: somente Situação Backoffice = Aprovado.
    # A normalização aceita diferenças de maiúsculas/minúsculas, acentos e espaços.
    analise_backoffice = analisar_situacao_backoffice(base)
    filtro_aprovado = analise_backoffice["mascara_aprovados"]

    filtro_periodo = datas.between(
        pd.Timestamp(inicio),
        pd.Timestamp(fim) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1),
        inclusive="both",
    )
    filtro = filtro_aprovado & atendimento.eq(TIPO) & filtro_periodo
    envio = base.loc[filtro, COLUNAS_ORIGEM].copy()

    # Lógica original: remove exatamente os dois últimos caracteres do Code Deep.
    envio["Code Deep"] = envio["Code Deep"].fillna("").astype(str).apply(lambda x: x[:-2] if len(x) >= 2 else "")
    envio["COMUNIDADE"] = envio[COMUNIDADES].replace(r"^\s*$", pd.NA, regex=True).bfill(axis=1).iloc[:, 0]
    envio.drop(columns=COMUNIDADES, inplace=True)
    envio["Data do registro"] = pd.to_datetime(envio["Data do registro"], errors="coerce", dayfirst=True).dt.strftime("%d/%m/%Y")
    envio.replace("NÃO SE APLICA", "", inplace=True)
    envio.fillna("", inplace=True)
    envio.rename(columns={"FOTO DO RG": "FOTO DO RG", "FOTO DO CPF:": "FOTO DO CPF"}, inplace=True)

    # A primeira URL representa a frente; a segunda URL representa o verso do RG.
    urls_rg = envio["FOTO DO RG"].apply(extrair_urls)
    envio["FOTO DO RG"] = urls_rg.apply(lambda links: links[0] if len(links) >= 1 else "")
    envio["FOTO DO RG - VERSO"] = urls_rg.apply(lambda links: links[1] if len(links) >= 2 else "")

    novos = envio["É novo cliente?"].astype(str).str.upper().str.strip()
    cadastro = envio.loc[novos.eq("NÃO")].copy()
    return cadastro, envio

def gerar_excel(cadastro, envio):
    # Mesmo layout e mesma ordem visual do arquivo manual enviado.
    ordem = [
        "Code Deep", "Data do registro", "ASRO", "COMUNIDADE", "CEP:",
        "Endereço", "Complemento:", "bairro", "É novo cliente?",
        "NOME COMPLETO:", "CPF:", "RG:", "FOTO DO RG", "FOTO DO RG - VERSO", "FOTO DO CPF",
        "DATA DE NASCIMENTO:", "TELEFONE PARA CONTATO:", "E-MAIL:",
        "POSSUI NIS?", "Informar o número do NIS:", "NÚMERO DO MEDIDOR:",
        "NÚMERO INSTALAÇÃO:", "UNIDADE CONSUMIDORA/ CÓDIGO DO CLIENTE:",
        "FAIXA DE ENQUADRAMENTO ESCOLHIDA PELO CLIENTE:",
        "DATA DE VENCIMENTO DA FATURA:"
    ]
    nomes_saida = {
        "Code Deep": "CODE DEEP",
        "Data do registro": "DATA DO REGISTRO",
        "ASRO": "ASRO",
        "COMUNIDADE": "COMUNIDADE",
        "CEP:": "CEP:",
        "Endereço": "ENDEREÇO",
        "Complemento:": "COMPLEMENTO:",
        "bairro": "BAIRRO",
        "É novo cliente?": "É NOVO CLIENTE?",
        "NOME COMPLETO:": "NOME COMPLETO:",
        "CPF:": "CPF:",
        "RG:": "RG:",
        "FOTO DO RG": "FOTO DO RG",
        "FOTO DO RG - VERSO": "FOTO DO RG - VERSO",
        "FOTO DO CPF": "FOTO DO CPF",
        "DATA DE NASCIMENTO:": "DATA DE NASCIMENTO:",
        "TELEFONE PARA CONTATO:": "TELEFONE PARA CONTATO:",
        "E-MAIL:": "E-MAIL:",
        "POSSUI NIS?": "POSSUI NIS?",
        "Informar o número do NIS:": "INFORMAR O NÚMERO DO NIS:",
        "NÚMERO DO MEDIDOR:": "NÚMERO DO MEDIDOR:",
        "NÚMERO INSTALAÇÃO:": "NÚMERO INSTALAÇÃO:",
        "UNIDADE CONSUMIDORA/ CÓDIGO DO CLIENTE:": "UNIDADE CONSUMIDORA/ CÓDIGO DO CLIENTE:",
        "FAIXA DE ENQUADRAMENTO ESCOLHIDA PELO CLIENTE:": "FAIXA DE ENQUADRAMENTO ESCOLHIDA PELO CLIENTE:",
        "DATA DE VENCIMENTO DA FATURA:": "DATA DE VENCIMENTO DA FATURA:",
    }
    larguras = [18.14, 18.43, 16.57, 30.71, 12.0, 66.71, 19.0, 13.71,
                16.57, 37.29, 15.43, 17.29, 25.0, 25.0, 23.71, 22.14, 25.14,
                40.71, 12.0, 29.14, 22.43, 22.0, 44.57, 51.0, 33.0]

    saidas = {
        "CADASTRO": cadastro.reindex(columns=ordem).rename(columns=nomes_saida),
        "PROJETOS ESPECIAIS": envio.reindex(columns=ordem).rename(columns=nomes_saida),
    }

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for aba, dados in saidas.items():
            dados.to_excel(writer, sheet_name=aba, index=False)
            ws = writer.book[aba]

            # Layout compacto do modelo manual: verde claro, fonte preta e altura comum.
            ws.freeze_panes = None
            ws.sheet_view.showGridLines = True
            ws.auto_filter.ref = ws.dimensions
            ws.row_dimensions[1].height = None

            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="63E6BE")
                cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

            for i, largura in enumerate(larguras, start=1):
                ws.column_dimensions[get_column_letter(i)].width = largura

            cab = {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}
            for coluna, rotulo in [("FOTO DO RG", "Abrir frente do RG"), ("FOTO DO RG - VERSO", "Abrir verso do RG"), ("FOTO DO CPF", "Abrir foto do CPF")]:
                idx = cab.get(coluna)
                if not idx:
                    continue
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(r, idx)
                    valor = str(cell.value).strip() if cell.value else ""
                    # Quando há mais de uma URL na mesma célula, usa a primeira como hyperlink.
                    link = re.split(r"[\r\n]+", valor)[0].strip()
                    if link.lower().startswith(("http://", "https://")):
                        cell.hyperlink = link
                        cell.value = rotulo
                        cell.style = "Hyperlink"

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Aptos Narrow", size=11, color="000000", underline="single" if cell.hyperlink else None)
                    cell.alignment = Alignment(vertical="center", wrap_text=False)

    out.seek(0)
    return out.getvalue()




# =========================================================
# ENCURTADOR DE URL
# =========================================================

ISGD_ENDPOINT = "https://is.gd/create.php"


def extrair_links_texto(valor):
    """Extrai URLs HTTP/HTTPS de textos e células, preservando a ordem."""
    if valor is None:
        return []
    texto = str(valor).replace("\r", "\n")
    encontrados = re.findall(r"https?://\S+", texto, flags=re.IGNORECASE)
    limpos = [url.rstrip(".,;:)]}") for url in encontrados]
    return list(dict.fromkeys(limpos))


def validar_url_para_encurtar(url):
    url = str(url).strip()
    if not re.match(r"^https?://", url, flags=re.IGNORECASE):
        raise ValueError("O endereço precisa começar com http:// ou https://.")
    if len(url) > 5000:
        raise ValueError("O endereço excede o limite de 5.000 caracteres.")
    return url


def encurtar_url(url, timeout=20):
    """Encurta uma URL real usando a API pública is.gd."""
    url = validar_url_para_encurtar(url)
    try:
        resposta = requests.get(
            ISGD_ENDPOINT,
            params={"format": "json", "url": url},
            headers={"User-Agent": "SistemaAutomacoes/1.0"},
            timeout=timeout,
        )
        resposta.raise_for_status()
        dados = resposta.json()
    except requests.Timeout as erro:
        raise RuntimeError("O serviço de encurtamento demorou para responder.") from erro
    except requests.RequestException as erro:
        raise RuntimeError(f"Não foi possível acessar o serviço de encurtamento: {erro}") from erro
    except ValueError as erro:
        raise RuntimeError("O serviço retornou uma resposta inválida.") from erro

    if dados.get("shorturl"):
        return dados["shorturl"]
    mensagem = dados.get("errormessage") or "O serviço não conseguiu encurtar este endereço."
    raise RuntimeError(mensagem)


def localizar_links_excel(uploaded_file):
    """Localiza links escritos e hiperlinks nativos em todas as abas do Excel."""
    from openpyxl import load_workbook
    uploaded_file.seek(0)
    wb = load_workbook(io.BytesIO(uploaded_file.read()), data_only=False)
    registros = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                links = []
                if cell.hyperlink and cell.hyperlink.target:
                    links.append(cell.hyperlink.target)
                links.extend(extrair_links_texto(cell.value))
                for ordem, link in enumerate(dict.fromkeys(links), start=1):
                    registros.append({
                        "ABA ORIGEM": ws.title,
                        "LINHA ORIGEM": cell.row,
                        "COLUNA ORIGEM": cell.column_letter,
                        "ORDEM DO LINK": ordem,
                        "URL ORIGINAL": link,
                    })
    return registros


def processar_excel_encurtador(uploaded_file):
    """Encurta links do Excel e gera uma planilha com origem, URL original e URL curta."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    registros = localizar_links_excel(uploaded_file)
    if not registros:
        raise ValueError("Nenhum link http ou https foi encontrado no arquivo.")

    cache = {}
    sucessos = 0
    for item in registros:
        original = item["URL ORIGINAL"]
        if original not in cache:
            try:
                cache[original] = (encurtar_url(original), "OK")
            except Exception as erro:
                cache[original] = ("", str(erro))
            time.sleep(0.15)
        curta, status = cache[original]
        item["URL ENCURTADA"] = curta
        item["STATUS"] = status
        if curta:
            sucessos += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "LINKS ENCURTADOS"
    headers = [
        "ABA ORIGEM", "LINHA ORIGEM", "COLUNA ORIGEM", "ORDEM DO LINK",
        "URL ORIGINAL", "URL ENCURTADA", "STATUS"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="63E6BE")
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center")

    for item in registros:
        ws.append([item[c] for c in headers])
        linha = ws.max_row
        if item["URL ORIGINAL"].lower().startswith(("http://", "https://")):
            ws.cell(linha, 5).hyperlink = item["URL ORIGINAL"]
            ws.cell(linha, 5).style = "Hyperlink"
        if item["URL ENCURTADA"]:
            ws.cell(linha, 6).hyperlink = item["URL ENCURTADA"]
            ws.cell(linha, 6).style = "Hyperlink"

    for i, width in enumerate([22, 16, 18, 16, 80, 32, 48], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return excel_bytes_from_wb(wb), len(registros), sucessos

# =========================================================
# CLASSIFICAÇÕES
# =========================================================

def classificar_periodo(hora):
    try:
        hora = int(hora)
    except Exception:
        return "FORA DO PERÍODO"

    if 7 <= hora <= 12:
        return "MANHÃ"

    if 13 <= hora <= 18:
        return "TARDE"

    return "FORA DO PERÍODO"


def classificar_faixa_horario_excel(hora):
    try:
        hora = int(hora)
    except Exception:
        return "FORA DO PERÍODO"

    if 7 <= hora <= 12:
        return "MANHÃ"

    if 13 <= hora <= 18:
        return "TARDE"

    return "FORA DO PERÍODO"


def classificar_tipo_atendimento(valor):
    v = normalize_text(valor)

    if "ADESAO" in v:
        return "ADESÕES"

    if "MORADOR AUSENTE" in v or "AUSENTE" in v:
        return "AUSENTES"

    if "RECUSA" in v:
        return "RECUSAS"

    if "AGENDAMENTO" in v or "AGEND" in v:
        return "AGENDAMENTOS"

    if "VAGO" in v:
        return "IMOVEIS VAGOS"

    if "DEMOLIDO" in v:
        return "DEMOLIDO"

    if "ABANDONADO" in v:
        return "ABANDONADO"

    return "OUTROS"


# =========================================================
# FILTRO DE ADESÕES
# =========================================================

def filtrar_adesoes_realizadas(df, col_tipo):
    if not col_tipo:
        return df.iloc[0:0].copy()

    tipo_norm = df[col_tipo].astype(str).apply(normalize_text)

    df_filtrado = df[
        tipo_norm.str.contains("ADESAO REALIZADA", na=False)
    ].copy()

    df_filtrado = df_filtrado.reset_index(drop=True)

    return df_filtrado


def gerar_mensagem_resultado_adesoes(df_adesoes):
    if df_adesoes is None or df_adesoes.empty:
        return "Resultado sobre adesões:\n\nNenhuma adesão encontrada no arquivo enviado."

    total_adesoes = len(df_adesoes)

    col_data = encontrar_coluna(df_adesoes, ["DATA", "REGISTRO"])
    col_asro = encontrar_coluna(df_adesoes, ["ASRO"])
    col_novo = encontrar_coluna(df_adesoes, ["NOVO"])

    data_texto = "Na data analisada"

    if col_data:
        datas = pd.to_datetime(
            df_adesoes[col_data],
            errors="coerce",
            dayfirst=True
        )

        datas_validas = datas.dropna()

        if not datas_validas.empty:
            data_min = datas_validas.min().strftime("%d/%m/%Y")
            data_max = datas_validas.max().strftime("%d/%m/%Y")

            if data_min == data_max:
                data_texto = f"No dia {data_min}"
            else:
                data_texto = f"No período de {data_min} a {data_max}"

    novos_clientes = 0

    if col_novo:
        novos_clientes = (
            df_adesoes[col_novo]
            .astype(str)
            .str.strip()
            .str.upper()
            .eq("SIM")
            .sum()
        )

    mensagem = "Resultado sobre adesões:\n\n"
    mensagem += (
        f"{data_texto} teve o total de {total_adesoes} adesões, "
        f"sendo {novos_clientes} novos clientes.\n\n"
    )

    if col_asro:
        for asro, dados_asro in sorted(df_adesoes.groupby(col_asro), key=lambda x: str(x[0])):
            total_asro = len(dados_asro)

            novos_asro = 0

            if col_novo:
                novos_asro = (
                    dados_asro[col_novo]
                    .astype(str)
                    .str.strip()
                    .str.upper()
                    .eq("SIM")
                    .sum()
                )

            mensagem += (
                f"{asro}: {total_asro} adesões, "
                f"sendo {novos_asro} novos clientes.\n"
            )

    return mensagem


def processar_filtro_adesoes(uploaded_file):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo

    df = read_excel_any(uploaded_file)
    df.columns = df.columns.astype(str).str.strip()

    col_tipo = encontrar_coluna(df, ["TIPO", "ATENDIMENTO"])

    if not col_tipo:
        raise ValueError("Coluna 'Tipo de atendimento' não encontrada.")

    df_adesoes = filtrar_adesoes_realizadas(df, col_tipo)

    df_final = pd.DataFrame()

    for coluna in df_adesoes.columns:
        nome = str(coluna).strip().lower()

        if "link backoffice" in nome:
            df_final["Link Backoffice"] = df_adesoes[coluna]

        elif "code deep" in nome:
            df_final["Code Deep"] = df_adesoes[coluna]

        elif "data do registro" in nome:
            df_final["Data do registro"] = df_adesoes[coluna]

        elif nome == "asro":
            df_final["ASRO"] = df_adesoes[coluna]

        elif nome == "nome completo:":
            df_final["Cliente"] = df_adesoes[coluna]

        elif nome in ["é novo cliente?", "e novo cliente?"]:
            df_final["É novo cliente?"] = df_adesoes[coluna]

        elif nome in ["situação backoffice", "situacao backoffice"]:
            df_final["Backoffice"] = df_adesoes[coluna]

        elif nome == "tipo de atendimento":
            df_final["Tipo de atendimento"] = df_adesoes[coluna]

    colunas_finais = [
        "Link Backoffice",
        "Code Deep",
        "Data do registro",
        "ASRO",
        "É novo cliente?",
        "Cliente",
        "Backoffice",
        "Tipo de atendimento"
    ]

    for col in colunas_finais:
        if col not in df_final.columns:
            df_final[col] = ""

    df_final = df_final[colunas_finais].reset_index(drop=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "RELATORIO"

    header_fill = PatternFill(start_color="4B0082", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_id, col_name in enumerate(df_final.columns, start=1):
        cell = ws.cell(row=1, column=col_id, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = borda

    for row_idx, row in enumerate(df_final.itertuples(index=False), start=2):
        for col_id, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_id, value=excel_value(value))
            cell.border = borda

    if len(df_final) > 0:
        ultima_linha = len(df_final) + 1

        tabela = Table(
            displayName="TabelaAdesoes",
            ref=f"A1:H{ultima_linha}"
        )

        estilo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False
        )

        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

    aplicar_bordas_e_larguras(ws)

    mensagem = gerar_mensagem_resultado_adesoes(df_adesoes)

    return (
        excel_bytes_from_wb(wb),
        f"Relatorio_Adesoes_{agora_sao_paulo().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mensagem
    )


# =========================================================
# TERMO DE DOAÇÃO
# =========================================================

def processar_termo_doacao(uploaded_file, logo_file=None):
    """
    Gera o Termo de Doação a partir do Excel bruto.

    Atualização aplicada somente nesta rotina:
    - Remove o uso de logo na geração do arquivo.
    - Mantém logo_file=None apenas por compatibilidade com chamadas antigas.
    - Corrige erro de colunas duplicadas, especialmente COMPLEMENTO.
    - Se houver colunas duplicadas com o mesmo nome, usa a primeira ocorrência.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df = read_excel_any(uploaded_file)
    df.columns = df.columns.astype(str).str.strip().str.upper()

    # =====================================================
    # FUNÇÕES INTERNAS
    # =====================================================

    def col_por_palavras(palavras):
        """
        Encontra a primeira coluna que contenha todas as palavras informadas.
        Exemplo: ["NOME", "COMPLETO"] encontra "NOME COMPLETO:".
        """
        return next(
            (c for c in df.columns if all(p in c for p in palavras)),
            None
        )

    def pegar_coluna_unica(df_base, nome_coluna):
        """
        Corrige erro quando existem colunas duplicadas no Excel.
        Se df_base.loc[:, nome_coluna] retornar várias colunas,
        usa somente a primeira.
        """
        if not nome_coluna:
            return ""

        dados = df_base.loc[:, nome_coluna]

        if isinstance(dados, pd.DataFrame):
            return dados.iloc[:, 0]

        return dados

    # =====================================================
    # IDENTIFICAR COLUNAS
    # =====================================================

    col_code = col_por_palavras(["CODE"])
    col_nome = col_por_palavras(["NOME", "COMPLETO"])
    col_novo = col_por_palavras(["NOVO"])
    col_asro = col_por_palavras(["ASRO"])
    col_end = col_por_palavras(["ENDERE"])
    col_comp = col_por_palavras(["COMPLEMENTO"])

    if not col_code or not col_nome or not col_novo:
        raise ValueError(
            "Colunas obrigatórias não encontradas: CODE, NOME COMPLETO e NOVO."
        )

    # =====================================================
    # FILTRAR SOMENTE NOVOS CLIENTES
    # =====================================================

    coluna_novo = (
        pegar_coluna_unica(df, col_novo)
        .astype(str)
        .str.upper()
        .str.strip()
    )

    df = df[coluna_novo == "SIM"].copy()

    # =====================================================
    # MONTAR DATAFRAME FINAL
    # =====================================================

    df_saida = pd.DataFrame()

    df_saida["CODE DEEP"] = pegar_coluna_unica(df, col_code)
    df_saida["NOME COMPLETO"] = pegar_coluna_unica(df, col_nome)

    if col_asro:
        df_saida["ASRO"] = pegar_coluna_unica(df, col_asro)
    else:
        df_saida["ASRO"] = ""

    if col_end:
        df_saida["ENDEREÇO"] = pegar_coluna_unica(df, col_end)
    else:
        df_saida["ENDEREÇO"] = ""

    # Correção principal: evita erro quando existem múltiplas colunas COMPLEMENTO.
    if col_comp:
        df_saida["COMPLEMENTO"] = pegar_coluna_unica(df, col_comp)
    else:
        df_saida["COMPLEMENTO"] = ""

    # =====================================================
    # TRATAMENTOS DE SEGURANÇA
    # =====================================================

    df_saida["CODE DEEP"] = df_saida["CODE DEEP"].fillna("").astype(str)
    df_saida["NOME COMPLETO"] = df_saida["NOME COMPLETO"].fillna("").astype(str)
    df_saida["ASRO"] = df_saida["ASRO"].fillna("SEM ASRO").astype(str).str.strip()
    df_saida["ENDEREÇO"] = df_saida["ENDEREÇO"].fillna("").astype(str)
    df_saida["COMPLEMENTO"] = df_saida["COMPLEMENTO"].fillna("").astype(str)

    # Remove duplicados antes de gerar o ranking e as abas do Termo de Doação.
    # Códigos preenchidos: mantém somente a primeira ocorrência de cada CODE DEEP.
    # Códigos vazios: remove somente linhas integralmente repetidas.
    df_saida = df_saida.drop_duplicates().copy()
    codigo_limpo = df_saida["CODE DEEP"].astype(str).str.strip()
    com_codigo = df_saida[codigo_limpo.ne("")].drop_duplicates(
        subset=["CODE DEEP"], keep="first"
    )
    sem_codigo = df_saida[codigo_limpo.eq("")]
    df_saida = pd.concat([com_codigo, sem_codigo], ignore_index=True)

    df_saida = df_saida.sort_values(by="NOME COMPLETO")

    # =====================================================
    # CRIAR EXCEL
    # =====================================================

    wb = Workbook()
    wb.remove(wb.active)

    periodo = agora_sao_paulo().strftime("%d-%m-%Y")

    # =====================================================
    # ABA RANKING
    # =====================================================

    ranking = (
        df_saida
        .groupby("ASRO")
        .size()
        .reset_index(name="TOTAL")
        .sort_values("TOTAL", ascending=False)
    )

    ws_rank = wb.create_sheet("RANKING")
    ws_rank.append(["ASRO", "TOTAL"])

    for _, r in ranking.iterrows():
        ws_rank.append([r["ASRO"], r["TOTAL"]])

    aplicar_bordas_e_larguras(ws_rank)

    # =====================================================
    # ESTILO
    # =====================================================

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    cor_cabecalho = PatternFill(start_color="0DB39E", fill_type="solid")
    fonte_cabecalho = Font(bold=True, color="FFFFFF")

    # =====================================================
    # ABAS POR ASRO
    # =====================================================

    for asro, dados in sorted(df_saida.groupby("ASRO"), key=lambda x: str(x[0])):
        ws = wb.create_sheet(title=safe_sheet_name(f"{asro} - {periodo}"))

        ws.merge_cells("A1:E1")
        ws["A1"] = "TERMO DE DOAÇÃO DE PADRÃO"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = cor_cabecalho
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:E2")
        ws["A2"] = f"Período: {periodo}"
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = [
            "CODE DEEP",
            "ASRO",
            "NOME COMPLETO",
            "ENDEREÇO",
            "COMPLEMENTO"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = fonte_cabecalho
            cell.fill = cor_cabecalho
            cell.alignment = Alignment(horizontal="center")
            cell.border = borda

        for i, row in enumerate(dados.itertuples(index=False), start=5):
            valores = [
                row[0],  # CODE DEEP
                row[2],  # ASRO
                row[1],  # NOME COMPLETO
                row[3],  # ENDEREÇO
                row[4],  # COMPLEMENTO
            ]

            for col, val in enumerate(valores, start=1):
                cell = ws.cell(row=i, column=col, value=excel_value(val))
                cell.border = borda

        aplicar_bordas_e_larguras(ws)

    return (
        excel_bytes_from_wb(wb),
        f"resultado_{agora_sao_paulo().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

# =========================================================
# RELATÓRIO DE AGENTES
# =========================================================

def processar_relatorio_agentes(uploaded_file):
    from openpyxl import Workbook

    df = read_excel_any(uploaded_file)
    df.columns = df.columns.astype(str).str.strip()

    col_asro = encontrar_coluna(df, ["ASRO"])
    col_agente = encontrar_coluna(df, ["NOME", "AGENTE"])
    col_tipo = encontrar_coluna(df, ["TIPO", "ATENDIMENTO"])
    col_data = encontrar_coluna(df, ["DATA"])

    faltando = []

    if not col_asro:
        faltando.append("ASRO")

    if not col_agente:
        faltando.append("Nome do agente")

    if not col_tipo:
        faltando.append("Tipo de atendimento")

    if not col_data:
        faltando.append("Data")

    if faltando:
        raise ValueError("Colunas obrigatórias não encontradas: " + ", ".join(faltando))

    df[col_data] = converter_data_brasil(df[col_data])
    df["TIPO_CLASS"] = df[col_tipo].apply(classificar_tipo_atendimento)

    wb = Workbook()
    wb.remove(wb.active)

    for asro, dados_asro in df.groupby(col_asro):
        ws = wb.create_sheet(safe_sheet_name(asro))

        tabela = (
            dados_asro
            .groupby([col_data, col_agente, "TIPO_CLASS"])
            .size()
            .reset_index(name="TOTAL")
        )

        ws.append(["DATA", "AGENTE", "TIPO", "TOTAL"])

        for _, r in tabela.iterrows():
            ws.append([
                formatar_data_brasil(r[col_data]),
                r[col_agente],
                r["TIPO_CLASS"],
                int(r["TOTAL"])
            ])

        aplicar_bordas_e_larguras(ws)

    return (
        excel_bytes_from_wb(wb),
        f"relatorio_agentes_{agora_sao_paulo().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


# =========================================================
# RELATÓRIO ENVIO / VISITAS / ADESÕES
# =========================================================

def preparar_base_ultima_visita(df, col_code, col_tipo):
    df = df.copy()

    df["IMOVEL"] = (
        df[col_code]
        .astype(str)
        .str.split("-")
        .str[0]
        .str.strip()
    )

    ordem = (
        df[col_code]
        .astype(str)
        .str.split("-")
        .str[1]
    )

    df["ORDEM_VISITA"] = pd.to_numeric(ordem, errors="coerce").fillna(0).astype(int)
    df["TIPO_CLASS"] = df[col_tipo].apply(classificar_tipo_atendimento)

    df = df.sort_values(by=["IMOVEL", "ORDEM_VISITA"])

    df_ultima_visita = df.drop_duplicates(
        subset=["IMOVEL"],
        keep="last"
    ).copy()

    return df, df_ultima_visita


def calcular_metricas_envio(base_visitas, base_ultima_visita, col_data):
    visitas_totais = len(base_visitas)
    imoveis_visitados = base_visitas["IMOVEL"].nunique()
    dias = base_visitas[col_data].dropna().nunique()
    media = round(visitas_totais / dias) if dias else 0

    adesoes = int((base_ultima_visita["TIPO_CLASS"] == "ADESÕES").sum())
    ausentes = int((base_ultima_visita["TIPO_CLASS"] == "AUSENTES").sum())
    recusas = int((base_ultima_visita["TIPO_CLASS"] == "RECUSAS").sum())
    agendamentos = int((base_ultima_visita["TIPO_CLASS"] == "AGENDAMENTOS").sum())

    def pct(valor):
        return valor / imoveis_visitados if imoveis_visitados else 0

    return {
        "Visitas totais": visitas_totais,
        "Imóveis visitados": imoveis_visitados,
        "Média visitas diárias": media,
        "Dias trabalhados": dias,
        "Moradores ausentes": ausentes,
        "% Moradores ausentes": pct(ausentes),
        "Adesões": adesoes,
        "% Adesões": pct(adesoes),
        "Recusas": recusas,
        "% Recusas": pct(recusas),
        "Agendamentos": agendamentos,
        "% Agendamentos": pct(agendamentos),
    }


def processar_relatorio_envio_visitas_adesoes(uploaded_file):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList

    df = read_excel_any(uploaded_file)
    df.columns = df.columns.astype(str).str.strip()

    col_asro = encontrar_coluna(df, ["ASRO"])
    col_tipo = encontrar_coluna(df, ["TIPO", "ATENDIMENTO"])
    col_agente = encontrar_coluna(df, ["NOME", "AGENTE"])
    col_data = encontrar_coluna(df, ["DATA", "REGISTRO"]) or encontrar_coluna(df, ["DATA"])
    col_code = encontrar_coluna(df, ["CODE"])

    faltando = []

    if not col_asro:
        faltando.append("ASRO")

    if not col_tipo:
        faltando.append("Tipo de atendimento")

    if not col_agente:
        faltando.append("Nome do agente")

    if not col_data:
        faltando.append("Data do registro")

    if not col_code:
        faltando.append("Code Deep")

    if faltando:
        raise ValueError("Colunas obrigatórias não encontradas: " + ", ".join(faltando))

    df[col_data] = converter_data_brasil(df[col_data])

    df_base, df_ultima = preparar_base_ultima_visita(df, col_code, col_tipo)

    wb = Workbook()
    wb.remove(wb.active)

    verde = PatternFill(start_color="006400", fill_type="solid")
    verde_medio = PatternFill(start_color="0DB39E", fill_type="solid")
    branco = Font(color="FFFFFF", bold=True)

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def criar_resumo(ws, titulo, base_visitas, base_ultima_visita):
        m = calcular_metricas_envio(base_visitas, base_ultima_visita, col_data)

        ws.append([titulo])
        ws.append([])
        ws.append(["Indicador", "Valor", "Percentual"])

        linhas = [
            ("Visitas totais", m["Visitas totais"], ""),
            ("Imóveis visitados", m["Imóveis visitados"], ""),
            ("Média visitas diárias", m["Média visitas diárias"], ""),
            ("Dias trabalhados", m["Dias trabalhados"], ""),
            ("Moradores ausentes", m["Moradores ausentes"], m["% Moradores ausentes"]),
            ("Adesões", m["Adesões"], m["% Adesões"]),
            ("Recusas", m["Recusas"], m["% Recusas"]),
            ("Agendamentos", m["Agendamentos"], m["% Agendamentos"]),
        ]

        for item in linhas:
            ws.append(list(item))

        ws.merge_cells("A1:C1")
        ws["A1"].fill = verde
        ws["A1"].font = branco
        ws["A1"].alignment = Alignment(horizontal="center")

        for cell in ws[3]:
            cell.fill = verde_medio
            cell.font = branco
            cell.alignment = Alignment(horizontal="center")

        for row in range(8, 12):
            ws.cell(row=row, column=3).number_format = "0.00%"

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = borda

        inicio = 14

        dados_grafico = pd.DataFrame([
            ["Adesões", m["Adesões"]],
            ["Moradores ausentes", m["Moradores ausentes"]],
            ["Recusas", m["Recusas"]],
            ["Agendamentos", m["Agendamentos"]],
        ], columns=["Indicador", "Total"])

        dados_grafico = dados_grafico.sort_values("Total", ascending=False)

        ws.cell(row=inicio, column=1, value="Indicador")
        ws.cell(row=inicio, column=2, value="Total")

        for idx, linha in enumerate(dados_grafico.itertuples(index=False), start=inicio + 1):
            ws.cell(row=idx, column=1, value=linha[0])
            ws.cell(row=idx, column=2, value=int(linha[1]))

        chart = BarChart()
        chart.title = titulo
        chart.y_axis.title = "Quantidade"
        chart.x_axis.title = "Indicadores"

        data_ref = Reference(ws, min_col=2, min_row=inicio, max_row=inicio + 4)
        cats_ref = Reference(ws, min_col=1, min_row=inicio + 1, max_row=inicio + 4)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True

        ws.add_chart(chart, "E3")

        aplicar_bordas_e_larguras(ws)

    ws_geral = wb.create_sheet("GERAL")

    criar_resumo(
        ws_geral,
        "RELATÓRIO DE ADESÕES, VISITAS E ÁREAS CORRELATAS",
        df_base,
        df_ultima
    )

    for asro, dados_asro in sorted(df_base.groupby(col_asro), key=lambda x: str(x[0])):
        ws = wb.create_sheet(safe_sheet_name(asro))

        imoveis_asro = dados_asro["IMOVEL"].unique()
        ultima_asro = df_ultima[df_ultima["IMOVEL"].isin(imoveis_asro)].copy()

        criar_resumo(
            ws,
            f"RELATÓRIO - {asro}",
            dados_asro,
            ultima_asro
        )

        linha = 22

        headers = [
            "Agente",
            "Visitas totais",
            "Imóveis visitados",
            "Adesões",
            "% Adesões",
            "Ausentes",
            "% Ausentes",
            "Recusas",
            "% Recusas",
            "Agendamentos",
            "% Agendamentos"
        ]

        for c, h in enumerate(headers, start=1):
            ws.cell(row=linha, column=c, value=h)

        linha += 1

        for agente, dados_agente in dados_asro.groupby(col_agente):
            imoveis_agente = dados_agente["IMOVEL"].unique()
            ultima_agente = ultima_asro[ultima_asro["IMOVEL"].isin(imoveis_agente)].copy()

            ma = calcular_metricas_envio(dados_agente, ultima_agente, col_data)

            valores = [
                excel_value(agente),
                ma["Visitas totais"],
                ma["Imóveis visitados"],
                ma["Adesões"],
                ma["% Adesões"],
                ma["Moradores ausentes"],
                ma["% Moradores ausentes"],
                ma["Recusas"],
                ma["% Recusas"],
                ma["Agendamentos"],
                ma["% Agendamentos"]
            ]

            for c, v in enumerate(valores, start=1):
                cell = ws.cell(row=linha, column=c, value=excel_value(v))

                if c in [5, 7, 9, 11]:
                    cell.number_format = "0.00%"

            linha += 1

        aplicar_bordas_e_larguras(ws)

    return (
        excel_bytes_from_wb(wb),
        f"Relatorio_Adesoes_Visitas_Areas_{agora_sao_paulo().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


# =========================================================
# ACOMPANHAMENTO DIÁRIO
# =========================================================

def preparar_acompanhamento(arquivos):
    bases = []

    for arquivo in arquivos:
        df = read_excel_any(arquivo)
        df.columns = df.columns.astype(str).str.strip()
        df["ARQUIVO_ORIGEM"] = arquivo.name
        bases.append(df)

    if not bases:
        raise ValueError("Nenhum arquivo enviado.")

    df = pd.concat(bases, ignore_index=True)

    col_asro = encontrar_coluna(df, ["ASRO"])
    col_agente = encontrar_coluna(df, ["NOME", "AGENTE"])
    col_data = encontrar_coluna(df, ["DATA", "REGISTRO"]) or encontrar_coluna(df, ["DATA"])
    col_horario = encontrar_coluna(df, ["HORARIO", "REGISTRO"]) or encontrar_coluna(df, ["HORA", "REGISTRO"])
    col_tipo = encontrar_coluna(df, ["TIPO", "ATENDIMENTO"])

    faltando = []

    if not col_asro:
        faltando.append("ASRO")

    if not col_agente:
        faltando.append("Nome do agente")

    if not col_data:
        faltando.append("Data do registro")

    if not col_horario:
        faltando.append("Horário de registro")

    if not col_tipo:
        faltando.append("Tipo de atendimento")

    if faltando:
        raise ValueError("Colunas obrigatórias não encontradas: " + ", ".join(faltando))

    df["DATA_REGISTRO_TRATADA"] = converter_data_brasil(df[col_data])

    horario_str = df[col_horario].astype(str).str.strip()

    hora_dt = pd.to_datetime(
        horario_str,
        errors="coerce",
        dayfirst=True
    ).dt.hour

    hora_num = pd.to_numeric(
        horario_str.str.extract(r"(\d{1,2})", expand=False),
        errors="coerce"
    )

    df["HORA_EXTRAIDA"] = hora_dt.fillna(hora_num).astype("Int64")

    df["PERIODO"] = df["HORA_EXTRAIDA"].apply(classificar_periodo)

    df["ASRO"] = df[col_asro].astype(str).str.strip()
    df["AGENTE"] = df[col_agente].astype(str).str.strip()
    df["TIPO_ORIGINAL"] = df[col_tipo].astype(str).str.strip()
    df["TIPO_CLASS"] = df[col_tipo].apply(classificar_tipo_atendimento)

    cols = {
        "ASRO": col_asro,
        "Agente": col_agente,
        "Data": col_data,
        "Horário": col_horario,
        "Tipo de atendimento": col_tipo
    }

    return df, cols


def resumo_agente_acomp(base, incluir_asro=True):
    colunas = [
        "Agente",
        "Visitas totais",
        "Imóveis visitados",
        "Adesões",
        "Ausentes",
        "Recusas",
        "Agendamentos",
        "Imoveis vagos"
    ]

    if incluir_asro:
        colunas = ["ASRO"] + colunas

    if base.empty:
        return pd.DataFrame(columns=colunas)

    grupos = ["ASRO", "AGENTE"] if incluir_asro else "AGENTE"

    linhas = []

    for chave, dados in base.groupby(grupos):
        if incluir_asro:
            asro, agente = chave
        else:
            asro = None
            agente = excel_value(chave)

        item = {
            "Agente": excel_value(agente),
            "Visitas totais": len(dados),
            "Imóveis visitados": len(dados),
            "Adesões": int((dados["TIPO_CLASS"] == "ADESÕES").sum()),
            "Ausentes": int((dados["TIPO_CLASS"] == "AUSENTES").sum()),
            "Recusas": int((dados["TIPO_CLASS"] == "RECUSAS").sum()),
            "Agendamentos": int((dados["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
            "Imoveis vagos": int(
                dados["TIPO_ORIGINAL"]
                .astype(str)
                .apply(lambda x: "VAGO" in normalize_text(x))
                .sum()
            ),
        }

        if incluir_asro:
            item = {
                "ASRO": excel_value(asro),
                **item
            }

        linhas.append(item)

    return pd.DataFrame(linhas).sort_values("Visitas totais", ascending=False)


def relatorio_final_simplificado(df_base):
    if df_base.empty:
        return pd.DataFrame(
            columns=[
                "ASRO",
                "Agente",
                "Período",
                "Horários",
                "Visitas",
                "Adesões",
                "Ausentes",
                "Recusas",
                "Agendamentos",
                "Imoveis vagos",
                "Principal atendimento"
            ]
        )

    linhas = []

    for (asro, agente, periodo), dados in df_base.groupby(["ASRO", "AGENTE", "PERIODO"]):
        tipos = dados["TIPO_ORIGINAL"].value_counts()
        principal = tipos.index[0] if len(tipos) else ""

        horarios = ", ".join(
            str(int(h))
            for h in sorted(dados["HORA_EXTRAIDA"].dropna().unique())
        )

        linhas.append({
            "ASRO": asro,
            "Agente": agente,
            "Período": periodo,
            "Horários": horarios,
            "Visitas": len(dados),
            "Adesões": int((dados["TIPO_CLASS"] == "ADESÕES").sum()),
            "Ausentes": int((dados["TIPO_CLASS"] == "AUSENTES").sum()),
            "Recusas": int((dados["TIPO_CLASS"] == "RECUSAS").sum()),
            "Agendamentos": int((dados["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
            "Imoveis vagos": int(
                dados["TIPO_ORIGINAL"]
                .astype(str)
                .apply(lambda x: "VAGO" in normalize_text(x))
                .sum()
            ),
            "Principal atendimento": principal,
        })

    return pd.DataFrame(linhas).sort_values(
        ["ASRO", "Período", "Visitas"],
        ascending=[True, True, False]
    )


def montar_resumo_geral_acompanhamento(df_base, hora_extracao=None):
    resumo_asro = (
        df_base
        .groupby("ASRO")
        .agg(
            Visitas=("AGENTE", "size"),
            Agentes=("AGENTE", "nunique")
        )
        .reset_index()
    )

    resumo_asro["Média por agente"] = resumo_asro.apply(
        lambda r: round(r["Visitas"] / r["Agentes"]) if r["Agentes"] else 0,
        axis=1
    )

    resumo_asro = resumo_asro[
        ["ASRO", "Visitas", "Média por agente"]
    ].sort_values("Visitas", ascending=False)

    ranking = (
        df_base
        .groupby(["AGENTE", "ASRO"])
        .size()
        .reset_index(name="Visitas")
        .sort_values("Visitas", ascending=False)
        .reset_index(drop=True)
    )

    ranking.insert(0, "RANKING", ranking.index + 1)

    datas_formatadas = []

    for data in sorted(df_base["DATA_REGISTRO_TRATADA"].dropna().unique()):
        datas_formatadas.append(formatar_data_brasil(data))

    if hora_extracao is None:
        hora_extracao = agora_sao_paulo().strftime("%H:%M:%S")

    return {
        "datas": ", ".join(datas_formatadas),
        "extracao": hora_extracao,
        "asros_atuando": df_base["ASRO"].nunique(),
        "total_registros": len(df_base),
        "total_agentes": df_base["AGENTE"].nunique(),
        "resumo_asro": resumo_asro,
        "ranking": ranking,
    }


def montar_resumo_agente_por_faixa(dados_agente):
    if dados_agente.empty:
        return pd.DataFrame(
            columns=[
                "Período",
                "Visitas",
                "Adesões",
                "Ausentes",
                "Recusas",
                "Agendamentos",
                "Imoveis vagos"
            ]
        )

    dados_agente = dados_agente.copy()
    dados_agente["FAIXA_EXCEL"] = dados_agente["HORA_EXTRAIDA"].apply(classificar_faixa_horario_excel)

    linhas = []

    for faixa in ["MANHÃ", "TARDE", "FORA DO PERÍODO"]:
        bloco = dados_agente[dados_agente["FAIXA_EXCEL"] == faixa]

        if bloco.empty:
            continue

        linhas.append({
            "Período": faixa,
            "Visitas": len(bloco),
            "Adesões": int((bloco["TIPO_CLASS"] == "ADESÕES").sum()),
            "Ausentes": int((bloco["TIPO_CLASS"] == "AUSENTES").sum()),
            "Recusas": int((bloco["TIPO_CLASS"] == "RECUSAS").sum()),
            "Agendamentos": int((bloco["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
            "Imoveis vagos": int(
                bloco["TIPO_ORIGINAL"]
                .astype(str)
                .apply(lambda x: "VAGO" in normalize_text(x))
                .sum()
            ),
        })

    linhas.append({
        "Período": "TOTAL",
        "Visitas": len(dados_agente),
        "Adesões": int((dados_agente["TIPO_CLASS"] == "ADESÕES").sum()),
        "Ausentes": int((dados_agente["TIPO_CLASS"] == "AUSENTES").sum()),
        "Recusas": int((dados_agente["TIPO_CLASS"] == "RECUSAS").sum()),
        "Agendamentos": int((dados_agente["TIPO_CLASS"] == "AGENDAMENTOS").sum()),
        "Imoveis vagos": int(
            dados_agente["TIPO_ORIGINAL"]
            .astype(str)
            .apply(lambda x: "VAGO" in normalize_text(x))
            .sum()
        ),
    })

    return pd.DataFrame(linhas)


def gerar_excel_acompanhamento(df_base):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    hora_extracao = agora_sao_paulo().strftime("%H:%M:%S")

    wb = Workbook()
    wb.remove(wb.active)

    verde = PatternFill(start_color="00B050", fill_type="solid")
    verde_escuro = PatternFill(start_color="006400", fill_type="solid")
    cinza = PatternFill(start_color="D9D9D9", fill_type="solid")
    cinza_claro = PatternFill(start_color="F2F2F2", fill_type="solid")

    branco = Font(color="FFFFFF", bold=True)
    fonte_titulo = Font(color="1F2937", bold=True, size=14)
    fonte_negrito = Font(bold=True)

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def estilizar_range(ws):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = borda
                    cell.alignment = Alignment(vertical="center")

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0

            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    def escrever_df(ws, df, row, col, header_fill=cinza):
        for j, nome_col in enumerate(df.columns, start=col):
            cell = ws.cell(row=row, column=j, value=str(nome_col))
            cell.fill = header_fill
            cell.font = fonte_negrito
            cell.border = borda
            cell.alignment = Alignment(horizontal="center")

        for i, linha in enumerate(df.itertuples(index=False), start=row + 1):
            for j, valor in enumerate(linha, start=col):
                cell = ws.cell(row=i, column=j, value=excel_value(valor))
                cell.border = borda
                cell.fill = cinza_claro if i % 2 == 0 else PatternFill(fill_type=None)

        return row + len(df) + 2

    geral = montar_resumo_geral_acompanhamento(
        df_base,
        hora_extracao=hora_extracao
    )

    ws = wb.create_sheet("GERAL")

    ws.merge_cells("A1:E1")
    ws["A1"] = "ACOMPANHAMENTO DIÁRIO - PRODUTIVIDADE DOS AGENTES"
    ws["A1"].fill = verde
    ws["A1"].font = fonte_titulo
    ws["A1"].alignment = Alignment(horizontal="center")

    labels = [
        "Data(s) filtrada(s)",
        "Horário da extração",
        "ASROs atuando",
        "Total de registros",
        "Total de agentes"
    ]

    valores = [
        geral["datas"],
        hora_extracao,
        geral["asros_atuando"],
        geral["total_registros"],
        geral["total_agentes"]
    ]

    for idx, label in enumerate(labels, start=3):
        ws.cell(row=idx, column=1, value=label).font = fonte_negrito

        cell_valor = ws.cell(row=idx, column=2, value=valores[idx - 3])

        if label == "Horário da extração":
            cell_valor.number_format = "@"

    resumo_start = 10

    ws.cell(row=resumo_start, column=1, value="ASRO").fill = cinza
    ws.cell(row=resumo_start, column=2, value="Visitas").fill = cinza
    ws.cell(row=resumo_start, column=3, value="Média por agente").fill = cinza

    for c in range(1, 4):
        ws.cell(row=resumo_start, column=c).font = fonte_negrito
        ws.cell(row=resumo_start, column=c).alignment = Alignment(horizontal="center")

    linha = resumo_start + 1

    for _, r in geral["resumo_asro"].iterrows():
        ws.cell(row=linha, column=1, value=r["ASRO"])
        ws.cell(row=linha, column=2, value=int(r["Visitas"]))
        ws.cell(row=linha, column=3, value=int(r["Média por agente"]))
        linha += 1

    ranking_start = 10

    for c, label in zip(
        range(5, 9),
        ["RANKING", "AGENTE", "ASRO", "Visitas"]
    ):
        ws.cell(row=ranking_start, column=c, value=label).fill = cinza
        ws.cell(row=ranking_start, column=c).font = fonte_negrito
        ws.cell(row=ranking_start, column=c).alignment = Alignment(horizontal="center")

    linha_rank = ranking_start + 1

    for _, r in geral["ranking"].iterrows():
        ws.cell(row=linha_rank, column=5, value=int(r["RANKING"]))
        ws.cell(row=linha_rank, column=6, value=r["AGENTE"])
        ws.cell(row=linha_rank, column=7, value=r["ASRO"])
        ws.cell(row=linha_rank, column=8, value=int(r["Visitas"]))
        linha_rank += 1

    estilizar_range(ws)

    for asro, dados_asro in df_base.groupby("ASRO"):
        ws_asro = wb.create_sheet(safe_sheet_name(asro))

        ws_asro.merge_cells("A1:G1")
        ws_asro["A1"] = f"ACOMPANHAMENTO DIÁRIO - ASRO {asro}"
        ws_asro["A1"].fill = verde
        ws_asro["A1"].font = fonte_titulo
        ws_asro["A1"].alignment = Alignment(horizontal="center")

        ws_asro.cell(row=3, column=1, value="Total de registros").font = fonte_negrito
        ws_asro.cell(row=3, column=2, value="Total de agentes").font = fonte_negrito
        ws_asro.cell(row=4, column=1, value=len(dados_asro))
        ws_asro.cell(row=4, column=2, value=dados_asro["AGENTE"].nunique())

        ws_asro.cell(
            row=6,
            column=1,
            value="* MANHÃ - 07H ÀS 12H | TARDE - 13H ÀS 18H | FORA DO PERÍODO"
        )
        ws_asro.cell(row=6, column=1).font = fonte_negrito

        linha = 8

        for agente, dados_agente in sorted(
            dados_asro.groupby("AGENTE"),
            key=lambda x: str(x[0])
        ):
            ws_asro.merge_cells(
                start_row=linha,
                start_column=1,
                end_row=linha,
                end_column=7
            )

            cell_agente = ws_asro.cell(row=linha, column=1, value=str(agente))
            cell_agente.fill = verde_escuro
            cell_agente.font = branco
            cell_agente.alignment = Alignment(horizontal="left")

            linha += 1

            tabela_agente = montar_resumo_agente_por_faixa(dados_agente)
            linha = escrever_df(ws_asro, tabela_agente, linha, 1, header_fill=cinza)
            linha += 1

        estilizar_range(ws_asro)

    ws_base = wb.create_sheet("BASE TRATADA")
    ws_base["A1"] = "BASE TRATADA"
    ws_base["A1"].fill = verde
    ws_base["A1"].font = fonte_titulo

    base_export = df_base[
        [
            "ARQUIVO_ORIGEM",
            "DATA_REGISTRO_TRATADA",
            "HORA_EXTRAIDA",
            "PERIODO",
            "ASRO",
            "AGENTE",
            "TIPO_ORIGINAL",
            "TIPO_CLASS"
        ]
    ].copy()

    escrever_df(ws_base, base_export, 3, 1, header_fill=cinza)
    estilizar_range(ws_base)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return (
        output.getvalue(),
        f"Acompanhamento_Diario_{agora_sao_paulo().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


def exibir_orientacao(objetivo, passos, resultado):
    """Exibe uma orientação azul e padronizada em todas as páginas."""
    passos_md = "\n".join(f"{i}. {passo}" for i, passo in enumerate(passos, start=1))
    st.info(
        f"""
**Para que serve**  
{objetivo}

**Como funciona**  
{passos_md}

**Resultado gerado**  
{resultado}
        """
    )


# =========================================================
# INTERFACE EM ABAS
# =========================================================

PAGINAS = [
    "Acompanhamento diário",
    "Filtro de Adesões",
    "Base de Envio Light",
    "Gerenciador de Hiperlinks",
    "Termo de Doação",
    "Relatório de Agentes",
    "Relatório Envio / Visitas / Adesões",
]
with st.sidebar:
    pagina = st.radio("Navegação", PAGINAS, label_visibility="collapsed")
    st.markdown("---")
    st.caption("Deep Field • Operação de campo")
    st.markdown(
        '<div class="deep-sidebar-footer notranslate" translate="no"><span>Desenvolvido por</span><strong>deep</strong><i class="deep-mark"></i></div>',
        unsafe_allow_html=True,
    )

st.markdown(
    """<div class="sca-hero notranslate" translate="no"><h1>Deep Field</h1><p>Automação e Inteligência Operacional</p></div>""",
    unsafe_allow_html=True,
)


# =========================================================
# ABA FILTRO DE ADESÕES
# =========================================================

if pagina == "Filtro de Adesões":
    st.header("Filtro de Adesões")

    exibir_orientacao(
        'Filtra o arquivo bruto e mantém somente as adesões realizadas, preparando os dados para conferência.',
        ['Envie o Excel bruto da operação.', 'Clique em **Executar Filtro de Adesões**.', 'Confira o resumo e baixe o Excel padronizado.'],
        'Excel de adesões e mensagem-resumo em TXT para comunicação operacional.',
    )

    
    arquivo = st.file_uploader(
        "Selecione o Excel bruto",
        type=["xlsx", "xls"],
        key="filtro_adesoes"
    )

    if arquivo and st.button("Executar Filtro de Adesões", key="btn_filtro"):
        try:
            data, nome, mensagem = processar_filtro_adesoes(arquivo)

            st.success("Relatório gerado com sucesso!")

            st.download_button(
                "Baixar Excel",
                data=data,
                file_name=nome,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.subheader("Mensagem automática para envio")

            st.text_area(
                "Copie a mensagem abaixo para enviar ao Caio:",
                value=mensagem,
                height=220
            )

            st.download_button(
                "Baixar mensagem em TXT",
                data=mensagem.encode("utf-8"),
                file_name=f"Mensagem_Adesoes_{agora_sao_paulo().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain"
            )

        except Exception as e:
            st.error(f"Erro: {e}")


# =========================================================
# ABA TERMO DE DOAÇÃO
# =========================================================

if pagina == "Base de Envio Light":
    st.header("Base de Envio Light")

    exibir_orientacao(
        'Cria a base de envio com apenas os registros aprovados pelo Backoffice.',
        ['Anexe a exportação bruta em Excel.', 'Confira o período identificado e selecione as datas desejadas.', 'Clique em **Gerar Base de Envio Light**.'],
        'Excel com as abas **CADASTRO** e **PROJETOS ESPECIAIS**, incluindo links de documentos.',
    )
    arquivo_light = st.file_uploader("Selecione a exportação bruta", type=["xlsx"], key="base_light_upload")
    if arquivo_light:
        try:
            conteudo_light = arquivo_light.getvalue()
            df_light, aba_light, header_light = ler_arquivo(conteudo_light)
            data_min_light, data_max_light = periodo_disponivel(df_light)
            analise_light = analisar_situacao_backoffice(df_light)
            st.caption(f"Aba: {aba_light} | Cabeçalho: linha {header_light + 1} | Período: {data_min_light:%d/%m/%Y} a {data_max_light:%d/%m/%Y}")
            l1, l2, l3 = st.columns(3)
            l1.metric("Registros", analise_light["total"])
            l2.metric("Aprovados", analise_light["aprovados"])
            l3.metric("Não aprovados excluídos", analise_light["nao_aprovados"])
            d1, d2 = st.columns(2)
            inicio_light = d1.date_input("Data inicial", value=data_min_light, format="DD/MM/YYYY", key="light_inicio")
            fim_light = d2.date_input("Data final", value=data_max_light, format="DD/MM/YYYY", key="light_fim")
            if inicio_light > fim_light:
                st.error("A data inicial não pode ser maior que a data final.")
            elif st.button("Gerar Base de Envio Light", key="btn_base_light", type="primary"):
                cadastro_light, projetos_light = processar(df_light, inicio_light, fim_light)
                arquivo_saida_light = gerar_excel(cadastro_light, projetos_light)
                st.success(f"Arquivo gerado: {len(projetos_light)} em PROJETOS ESPECIAIS e {len(cadastro_light)} em CADASTRO.")
                st.download_button(
                    "Baixar Base de Envio Light",
                    data=arquivo_saida_light,
                    file_name=f"base_envio_light_aprovados_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"Erro: {e}")

if pagina == "Gerenciador de Hiperlinks":
    st.header("Gerenciador de Hiperlinks")

    exibir_orientacao(
        'Transforma URLs extensas em textos clicáveis e, opcionalmente, encurta endereços externos.',
        ['Envie um Excel com a coluna de URL e a coluna do nome desejado.', 'Selecione as duas colunas correspondentes.', 'Gere e baixe o arquivo com hiperlinks amigáveis.'],
        'Excel com textos clicáveis; o encurtamento externo permanece disponível como opção secundária.',
    )

    
    st.warning(
        "Os endereços são enviados ao serviço público is.gd para encurtamento. "
        "Não envie links internos, confidenciais ou que contenham dados sensíveis sem autorização."
    )

    st.subheader("Modo recomendado: texto amigável")
    st.caption("Transforma uma URL extensa em um texto clicável escolhido pelo usuário, sem enviar dados a serviços externos.")
    arquivo_amigavel = st.file_uploader("Selecione o Excel com as colunas de URL e nome", type=["xlsx"], key="excel_hiperlink_amigavel")
    if arquivo_amigavel:
        try:
            previa_links = read_excel_any(arquivo_amigavel, dtype=str)
            colunas_links = list(previa_links.columns)
            c_url, c_nome = st.columns(2)
            coluna_url = c_url.selectbox("Coluna com a URL", colunas_links, key="col_url_amigavel")
            coluna_nome = c_nome.selectbox("Coluna com o texto desejado", colunas_links, index=min(1, len(colunas_links)-1), key="col_nome_amigavel")
            if st.button("Gerar Excel com hiperlinks amigáveis", type="primary", key="btn_hiperlink_amigavel"):
                from openpyxl import load_workbook
                arquivo_amigavel.seek(0)
                wb_h = load_workbook(io.BytesIO(arquivo_amigavel.read()))
                ws_h = wb_h.active
                cab_h = {str(c.value).strip(): c.column for c in ws_h[1] if c.value is not None}
                idx_url, idx_nome = cab_h.get(coluna_url), cab_h.get(coluna_nome)
                if not idx_url or not idx_nome:
                    raise ValueError("As colunas selecionadas não foram encontradas na primeira linha da aba ativa.")
                for linha_h in range(2, ws_h.max_row + 1):
                    url_h = str(ws_h.cell(linha_h, idx_url).value or "").strip()
                    nome_h = str(ws_h.cell(linha_h, idx_nome).value or "").strip()
                    if url_h.lower().startswith(("http://", "https://")):
                        cel_h = ws_h.cell(linha_h, idx_nome)
                        cel_h.value = nome_h or "Abrir link"
                        cel_h.hyperlink = url_h
                        cel_h.style = "Hyperlink"
                out_h = io.BytesIO(); wb_h.save(out_h); out_h.seek(0)
                st.download_button("Baixar Excel com hiperlinks", out_h.getvalue(), "hiperlinks_amigaveis.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
        except Exception as erro:
            st.error(f"Erro no modo de texto amigável: {erro}")

    with st.expander("Modo opcional: encurtar URL por serviço externo"):
        st.caption("Este modo usa o serviço público is.gd e depende de acesso à internet.")

    arquivo_links = st.file_uploader(
        "Selecione o Excel com links",
        type=["xlsx"],
        key="excel_encurtador_final",
    )

    if arquivo_links:
        try:
            links_localizados = localizar_links_excel(arquivo_links)
            st.metric("Links identificados", len(links_localizados))

            if not links_localizados:
                st.warning("Nenhum link http ou https foi encontrado no arquivo.")
            elif st.button(
                "Encurtar links e gerar Excel",
                key="btn_excel_encurtador_final",
                type="primary",
                use_container_width=True,
            ):
                with st.spinner("Localizando, separando e encurtando os links..."):
                    resultado, total, sucessos = processar_excel_encurtador(arquivo_links)

                st.success(
                    f"Processamento concluído: {sucessos} de {total} links foram encurtados."
                )

                if sucessos < total:
                    st.warning(
                        "Alguns links não foram encurtados. Consulte a coluna STATUS no arquivo gerado."
                    )

                st.download_button(
                    "Baixar Excel com links encurtados",
                    data=resultado,
                    file_name=f"links_encurtados_{agora_sao_paulo():%Y%m%d_%H%M%S}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
        except Exception as erro:
            st.error(f"Erro: {erro}")


if pagina == "Termo de Doação":
    st.header("Termo de Doação")

    exibir_orientacao(
        'Organiza os novos clientes e gera o arquivo semanal do Termo de Doação por ASRO.',
        ['Envie o Excel bruto.', 'Clique em **Gerar Termo de Doação**.', 'Baixe o arquivo organizado por comunidade e agente.'],
        'Excel do Termo de Doação com ranking e abas separadas por ASRO.',
    )

    
    arquivo = st.file_uploader(
        "Selecione o Excel bruto",
        type=["xlsx"],
        key="termo_doacao"
    )

    if arquivo and st.button("Gerar Termo de Doação", key="btn_termo"):
        try:
            data, nome = processar_termo_doacao(arquivo)

            st.success("Termo gerado com sucesso!")

            st.download_button(
                "Baixar Excel",
                data=data,
                file_name=nome,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Erro: {e}")

# =========================================================
# ABA RELATÓRIO DE AGENTES
# =========================================================

if pagina == "Relatório de Agentes":
    st.header("Relatório Semanal dos Agentes")

    exibir_orientacao(
        'Consolida a produtividade dos agentes por ASRO, data e tipo de atendimento.',
        ['Envie o Excel bruto da operação.', 'Clique em **Gerar Relatório de Agentes**.', 'Baixe o relatório para análise da equipe.'],
        'Excel segmentado por ASRO com a produção individual dos agentes.',
    )

    
    arquivo = st.file_uploader(
        "Selecione o Excel bruto",
        type=["xlsx", "xls"],
        key="rel_agentes"
    )

    if arquivo and st.button("Gerar Relatório de Agentes", key="btn_agentes"):
        try:
            data, nome = processar_relatorio_agentes(arquivo)

            st.success("Relatório gerado com sucesso!")

            st.download_button(
                "Baixar Excel",
                data=data,
                file_name=nome,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Erro: {e}")


# =========================================================
# ABA RELATÓRIO ENVIO / VISITAS / ADESÕES
# =========================================================

if pagina == "Relatório Envio / Visitas / Adesões":
    st.header("Relatório Envio / Visitas / Adesões")

    exibir_orientacao(
        'Calcula visitas, imóveis únicos e o resultado da última visita de cada imóvel.',
        ['Envie o Excel bruto.', 'O sistema identifica Code Deep, agente, ASRO, data e tipo de atendimento.', 'Gere e baixe o relatório consolidado.'],
        'Excel com resumo geral, páginas por ASRO, indicadores e gráficos.',
    )

    
    arquivo = st.file_uploader(
        "Selecione o Excel bruto",
        type=["xlsx", "xls"],
        key="rel_envio"
    )

    if arquivo and st.button("Gerar Relatório Envio / Visitas / Adesões", key="btn_envio"):
        try:
            data, nome = processar_relatorio_envio_visitas_adesoes(arquivo)

            st.success("Relatório gerado com sucesso!")

            st.download_button(
                "Baixar Excel",
                data=data,
                file_name=nome,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Erro: {e}")


# =========================================================
# ABA ACOMPANHAMENTO DIÁRIO
# =========================================================

if pagina == "Acompanhamento diário":
    st.header("Acompanhamento diário")

    exibir_orientacao(
        'Acompanha a sincronização e a produtividade diária dos 17 agentes cadastrados.',
        ['Envie um ou vários arquivos brutos.', 'Filtre comunidade e data.', 'Analise visitas, adesões, horários e agentes sem registros.', 'Gere Excel, PDFs por comunidade e PowerPoint.'],
        'Dashboard interativo e arquivos em Excel, PDF e PowerPoint com os dados filtrados.',
    )

    
    arquivos = st.file_uploader(
        "Enviar arquivo bruto ou vários arquivos",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="upload_acompanhamento"
    )

    if arquivos:
        try:
            df_original, cols = preparar_acompanhamento(arquivos)

            st.success(
                f"Arquivos carregados: {len(arquivos)} | "
                f"Total de registros: {len(df_original)}"
            )

            st.caption(
                "Colunas detectadas: "
                + " | ".join([f"{k}: {v}" for k, v in cols.items()])
            )

            st.subheader("Filtros")

            f1, f2 = st.columns(2)

            asros = sorted(
                df_original["ASRO"]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )

            datas = sorted(
                df_original["DATA_REGISTRO_TRATADA"]
                .dropna()
                .unique()
                .tolist()
            )

            with f1:
                asros_sel = st.multiselect(
                    "ASRO",
                    asros,
                    default=asros
                )

            with f2:
                datas_sel = st.multiselect(
                    "Data",
                    datas,
                    default=datas
                )

            df_dash = df_original[
                df_original["ASRO"].isin(asros_sel)
                & df_original["DATA_REGISTRO_TRATADA"].isin(datas_sel)
            ].copy()

            if df_dash.empty:
                st.warning("Nenhum dado encontrado com os filtros selecionados.")

            else:
                resumo_cadastro = resumo_operacional_com_cadastro(df_dash)
                agentes_com_registro = int((resumo_cadastro["Visitas"] > 0).sum())
                agentes_sem_registro = int((resumo_cadastro["Visitas"] == 0).sum())
                st.subheader("Controle de sincronização dos agentes")
                s1, s2, s3 = st.columns(3)
                s1.metric("Agentes cadastrados", len(resumo_cadastro))
                s2.metric("Com registros", agentes_com_registro)
                s3.metric("Sem registros", agentes_sem_registro)
                produtivos = resumo_cadastro[resumo_cadastro["Visitas"] > 0].copy()
                if not produtivos.empty:
                    fig_agentes = px.pie(
                        produtivos, names="AGENTE", values="Visitas", hole=0.28,
                        color_discrete_sequence=["#075E54", "#0B7368", "#0B8F7A", "#13A88E", "#35BDB1", "#68CEC5", "#9BE0D9"],
                        title="Participação de cada agente no total de visitas",
                    )
                    fig_agentes.update_traces(textposition="outside", textinfo="label+value+percent")
                    fig_agentes.update_layout(height=520, margin=dict(l=20,r=20,t=70,b=20), legend_title_text="Agentes")
                    st.plotly_chart(fig_agentes, use_container_width=True)
                sem_nomes = resumo_cadastro.loc[resumo_cadastro["Visitas"] == 0, ["ASRO", "AGENTE"]]
                if not sem_nomes.empty:
                    st.warning("Agentes cadastrados sem registros no período selecionado")
                    st.dataframe(sem_nomes, use_container_width=True, hide_index=True)
                st.dataframe(
                    resumo_cadastro.rename(columns={"ASRO":"Comunidade", "AGENTE":"Agente", "Adesoes":"Adesões", "Imoveis_vagos":"Imóveis vagos"}),
                    use_container_width=True, hide_index=True,
                    column_config={"Taxa de adesão": st.column_config.NumberColumn(format="%.1f%%")},
                )

                total_visitas = len(df_dash)
                agentes_unicos = df_dash["AGENTE"].nunique()

                total_manha = int((df_dash["PERIODO"] == "MANHÃ").sum())
                total_tarde = int((df_dash["PERIODO"] == "TARDE").sum())
                total_fora = int((df_dash["PERIODO"] == "FORA DO PERÍODO").sum())

                k1, k2, k3, k4, k5 = st.columns(5)

                k1.metric("Total de registros", total_visitas)
                k2.metric("Manhã", total_manha)
                k3.metric("Tarde", total_tarde)
                k4.metric("Fora do período", total_fora)
                k5.metric(
                    "Média/agente",
                    round(total_visitas / agentes_unicos) if agentes_unicos else 0
                )

                st.subheader("Painel")

                col_pizza, col_top = st.columns([1.25, 2])

                with col_pizza:
                    st.markdown("**Distribuição por período**")

                    periodo = (
                        df_dash
                        .groupby("PERIODO")
                        .size()
                        .reset_index(name="TOTAL")
                    )

                    fig_pie = px.pie(
                        periodo,
                        names="PERIODO",
                        values="TOTAL",
                        hole=0.35,
                        color_discrete_sequence=[
                            "#38bdf8",
                            "#22c55e",
                            "#94a3b8"
                        ]
                    )

                    fig_pie.update_traces(
                        textposition="inside",
                        textinfo="percent+label+value"
                    )

                    fig_pie.update_layout(
                        height=430,
                        margin=dict(l=10, r=10, t=40, b=10),
                        showlegend=True
                    )

                    st.plotly_chart(fig_pie, use_container_width=True)

                with col_top:
                    st.markdown("**Top agentes por visitas**")

                    top_agentes = resumo_agente_acomp(
                        df_dash,
                        incluir_asro=True
                    ).head(15)

                    fig_bar = px.bar(
                        top_agentes,
                        x="Visitas totais",
                        y="Agente",
                        color="ASRO",
                        orientation="h",
                        text="Visitas totais"
                    )

                    fig_bar.update_layout(
                        height=430,
                        yaxis={"categoryorder": "total ascending"}
                    )

                    st.plotly_chart(fig_bar, use_container_width=True)

                st.subheader("Divisão por período, ASRO e agente")

                for periodo_nome in ["MANHÃ", "TARDE", "FORA DO PERÍODO"]:
                    dados_periodo = df_dash[df_dash["PERIODO"] == periodo_nome]

                    st.markdown(f"### {periodo_nome}")

                    if dados_periodo.empty:
                        st.info("Sem registros nesse período.")
                        continue

                    for asro_nome in sorted(
                        dados_periodo["ASRO"]
                        .dropna()
                        .astype(str)
                        .unique()
                        .tolist()
                    ):
                        dados_asro_periodo = dados_periodo[
                            dados_periodo["ASRO"] == asro_nome
                        ]

                        with st.expander(
                            f"ASRO {asro_nome} - {periodo_nome}",
                            expanded=True
                        ):
                            heat = (
                                dados_asro_periodo
                                .groupby(["AGENTE", "HORA_EXTRAIDA"])
                                .size()
                                .reset_index(name="VISITAS")
                            )

                            pivot = heat.pivot_table(
                                index="AGENTE",
                                columns="HORA_EXTRAIDA",
                                values="VISITAS",
                                fill_value=0
                            )

                            if not pivot.empty:
                                pivot = pivot.reindex(sorted(pivot.columns), axis=1)

                                fig_heat = px.imshow(
                                    pivot,
                                    text_auto=True,
                                    aspect="auto",
                                    color_continuous_scale=[
                                        "#f7fbff",
                                        "#c6dbef",
                                        "#6baed6",
                                        "#2171b5",
                                        "#08306b"
                                    ],
                                    labels=dict(
                                        x="Horário",
                                        y="Agente",
                                        color="Visitas"
                                    )
                                )

                                fig_heat.update_layout(
                                    height=max(280, 28 * len(pivot.index)),
                                    margin=dict(l=20, r=20, t=30, b=20)
                                )

                                st.plotly_chart(fig_heat, use_container_width=True)

                            st.dataframe(
                                resumo_agente_acomp(
                                    dados_asro_periodo,
                                    incluir_asro=False
                                ),
                                use_container_width=True
                            )

                st.subheader("Relatório final detalhado")

                st.dataframe(
                    relatorio_final_simplificado(df_dash),
                    use_container_width=True
                )

                st.subheader("Gerar arquivo Excel")

                if st.button("Gerar relatório final em Excel", key="btn_excel_acompanhamento"):
                    excel_data, nome_arquivo = gerar_excel_acompanhamento(df_dash)

                    st.success("Arquivo Excel gerado com sucesso!")

                    st.download_button(
                        "Baixar Excel do acompanhamento diário",
                        data=excel_data,
                        file_name=nome_arquivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                st.subheader("Relatórios em PDF e PowerPoint")
                periodo_pdf = ", ".join(formatar_data_brasil(d) for d in sorted(df_dash["DATA_REGISTRO_TRATADA"].dropna().unique())) or "Período selecionado"

                diagnostico_pdf = diagnostico_relatorio(df_dash)
                d1, d2, d3 = st.columns(3)
                d1.metric("Registros filtrados", diagnostico_pdf["total_filtrado"])
                d2.metric("Associados às comunidades", diagnostico_pdf["total_associado"])
                d3.metric("Não associados", diagnostico_pdf["total_nao_associado"])

                st.caption(
                    "Registros usados nos PDFs: "
                    + " | ".join(
                        f"{comunidade.title()}: {total}"
                        for comunidade, total in diagnostico_pdf["por_comunidade"].items()
                    )
                )
                if diagnostico_pdf["asros_nao_reconhecidas"]:
                    st.warning(
                        "Valores de ASRO não reconhecidos e não associados por agente: "
                        + ", ".join(diagnostico_pdf["asros_nao_reconhecidas"])
                    )
                if diagnostico_pdf["agentes_nao_cadastrados"]:
                    st.info(
                        "Agentes encontrados no Excel e incluídos no relatório, mas ainda não cadastrados: "
                        + ", ".join(diagnostico_pdf["agentes_nao_cadastrados"])
                    )

                if st.button("Preparar relatórios para download", key="btn_relatorios_sca", type="primary"):
                    if diagnostico_pdf["total_associado"] == 0:
                        st.error(
                            "Nenhum registro filtrado foi associado a Salgueiro, Corte Oito ou Chacrinha. "
                            "Confira os valores da coluna ASRO e os nomes dos agentes."
                        )
                    else:
                        mapa_pdf = {
                            "pdf_salgueiro": "SALGUEIRO",
                            "pdf_corte_oito": "CORTE OITO",
                            "pdf_chacrinha": "CHACRINHA",
                        }
                        for chave_pdf, comunidade_pdf in mapa_pdf.items():
                            if diagnostico_pdf["por_comunidade"][comunidade_pdf] > 0:
                                st.session_state[chave_pdf] = gerar_pdf_comunidade(
                                    df_dash, comunidade_pdf, LOGO_LIGHT, periodo_pdf
                                )
                            else:
                                st.session_state.pop(chave_pdf, None)
                        st.session_state["pdf_mensal"] = gerar_pdf_mensal(df_dash, LOGO_LIGHT, periodo_pdf)
                        st.session_state["ppt_sca"] = gerar_powerpoint_resumo(df_dash, LOGO_LIGHT, periodo_pdf)
                        st.success(
                            f"Relatórios preparados com {diagnostico_pdf['total_associado']} registros filtrados."
                        )

                b1,b2,b3,b4 = st.columns(4)
                if "pdf_salgueiro" in st.session_state:
                    b1.download_button("PDF Salgueiro", st.session_state["pdf_salgueiro"], "Deep_Field_Salgueiro.pdf", "application/pdf", use_container_width=True)
                if "pdf_corte_oito" in st.session_state:
                    b2.download_button("PDF Corte Oito", st.session_state["pdf_corte_oito"], "Deep_Field_Corte_Oito.pdf", "application/pdf", use_container_width=True)
                if "pdf_chacrinha" in st.session_state:
                    b3.download_button("PDF Chacrinha", st.session_state["pdf_chacrinha"], "Deep_Field_Chacrinha.pdf", "application/pdf", use_container_width=True)
                if "pdf_mensal" in st.session_state:
                    b4.download_button("PDF consolidado", st.session_state["pdf_mensal"], "Deep_Field_Relatorio_Consolidado.pdf", "application/pdf", use_container_width=True)
                if "ppt_sca" in st.session_state:
                    st.download_button("Baixar PowerPoint executivo", st.session_state["ppt_sca"], "Deep_Field_Resultados_Operacionais.pptx", "application/vnd.openxmlformats-officedocument.presentationml.presentation", use_container_width=True)

        except Exception as e:
            st.error(f"Erro ao gerar acompanhamento diário: {e}")
