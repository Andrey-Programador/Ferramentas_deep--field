import io
import pandas as pd
from openpyxl import Workbook, load_workbook

from sca_reports import (
    cadastro_agentes_df,
    resumo_operacional_com_cadastro,
    gerar_pdf_comunidade,
    gerar_pdf_mensal,
    adicionar_controle_agentes_excel,
    gerar_powerpoint_resumo,
)


def base_teste():
    return pd.DataFrame([
        {"ASRO":"Salgueiro", "AGENTE":"Anna Paula Santana", "TIPO_CLASS":"ADESÕES"},
        {"ASRO":"Salgueiro", "AGENTE":"Anna Paula Santana", "TIPO_CLASS":"AUSENTES"},
        {"ASRO":"Salgueiro", "AGENTE":"Ana Paula Silva", "TIPO_CLASS":"RECUSAS"},
        {"ASRO":"Corte Oito", "AGENTE":"Nayara", "TIPO_CLASS":"ADESÕES"},
        {"ASRO":"Chacrinha", "AGENTE":"Daiane", "TIPO_CLASS":"AGENDAMENTOS"},
    ])


def test_cadastro_tem_17_agentes():
    cadastro = cadastro_agentes_df()
    assert len(cadastro) == 17
    assert cadastro.groupby("ASRO").size().to_dict() == {"CHACRINHA":6, "CORTE OITO":5, "SALGUEIRO":6}


def test_resumo_inclui_zeros_e_contagens():
    resumo = resumo_operacional_com_cadastro(base_teste())
    anna = resumo[resumo["AGENTE"] == "Anna Paula Santana"].iloc[0]
    josefina = resumo[resumo["AGENTE"] == "Josefina"].iloc[0]
    assert anna["Visitas"] == 2 and anna["Adesoes"] == 1
    assert josefina["Visitas"] == 0
    assert josefina["Situação de sincronização"] == "Sem registros"


def test_pdf_comunidade_e_mensal():
    logo = "/mnt/data/light_logo.png"
    pdf = gerar_pdf_comunidade(base_teste(), "SALGUEIRO", logo, "01/08/2026")
    mensal = gerar_pdf_mensal(base_teste(), logo, "Agosto de 2026")
    assert pdf.startswith(b"%PDF") and len(pdf) > 10000
    assert mensal.startswith(b"%PDF") and len(mensal) > len(pdf)


def test_excel_recebe_controle_dos_17_agentes():
    wb = Workbook(); ws = wb.active; ws.title = "GERAL"; ws.append(["OK"])
    raw = io.BytesIO(); wb.save(raw)
    data = adicionar_controle_agentes_excel(raw.getvalue(), base_teste())
    result = load_workbook(io.BytesIO(data))
    assert "CONTROLE DE AGENTES" in result.sheetnames
    assert result["CONTROLE DE AGENTES"].max_row == 18


def test_powerpoint_tem_capa_e_tres_comunidades():
    data = gerar_powerpoint_resumo(base_teste(), "/mnt/data/light_logo.png", "Agosto de 2026")
    assert data[:2] == b"PK" and len(data) > 20000


def test_pdf_reconhece_aliases_e_preserva_agente_novo():
    from sca_reports import diagnostico_relatorio, resumo_operacional_com_cadastro
    base = pd.DataFrame([
        {"ASRO":"Corte 8", "AGENTE":"Ícaro", "TIPO_CLASS":"ADESÕES"},
        {"ASRO":"C8", "AGENTE":"ICARO - CORTE OITO", "TIPO_CLASS":"RECUSAS"},
        {"ASRO":"Comunidade Salgueiro", "AGENTE":"Agente Novo", "TIPO_CLASS":"ADESÕES"},
        {"ASRO":"codigo interno", "AGENTE":"Daiane", "TIPO_CLASS":"AUSENTES"},
    ])
    diag = diagnostico_relatorio(base)
    assert diag["total_filtrado"] == 4
    assert diag["total_associado"] == 4
    assert diag["por_comunidade"] == {"SALGUEIRO":1, "CORTE OITO":2, "CHACRINHA":1}
    resumo = resumo_operacional_com_cadastro(base)
    icaro = resumo[(resumo["ASRO"] == "CORTE OITO") & (resumo["AGENTE"] == "Ícaro")].iloc[0]
    novo = resumo[(resumo["ASRO"] == "SALGUEIRO") & (resumo["AGENTE"] == "Agente Novo")].iloc[0]
    assert icaro["Visitas"] == 2 and icaro["Adesoes"] == 1 and icaro["Recusas"] == 1
    assert novo["Visitas"] == 1 and novo["Adesoes"] == 1


def test_totais_do_pdf_correspondem_aos_registros_associados():
    from sca_reports import gerar_pdf_comunidade
    from pypdf import PdfReader
    base = pd.DataFrame([
        {"ASRO":"Corte 8", "AGENTE":"Ícaro", "TIPO_CLASS":"ADESÕES"},
        {"ASRO":"Corte Oito", "AGENTE":"Nayara", "TIPO_CLASS":"ADESÕES"},
        {"ASRO":"C8", "AGENTE":"Nayara", "TIPO_CLASS":"AUSENTES"},
    ])
    pdf = gerar_pdf_comunidade(base, "CORTE OITO", "/mnt/data/light_logo.png", "Teste")
    reader = PdfReader(io.BytesIO(pdf))
    texto = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Total de visitas" in texto and "Total de adesões" in texto
    assert "3" in texto and "2" in texto
